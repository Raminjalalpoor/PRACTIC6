import tempfile
import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from app import create_app


CSV = """Дата;Товар;Количество;Цена;Категория
01.01.2025;Молоко;10;89,5;Молочные продукты
01.02.2025;Молоко;12;90;Молочные продукты
01.03.2025;Молоко;14;92;Молочные продукты
"""


class AppTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.uploads = Path(self.temp_dir.name) / "uploads"
        self.uploads.mkdir()
        self.app = create_app(
            {
                "TESTING": True,
                "SECRET_KEY": "test-key",
                "UPLOAD_FOLDER": self.uploads,
                "DATA_FOLDER": Path(__file__).parents[1] / "data",
            }
        )
        self.client = self.app.test_client()

    def tearDown(self):
        self.temp_dir.cleanup()

    def upload(self, content=CSV, filename="sales.csv", follow_redirects=False):
        return self.client.post(
            "/upload",
            data={"sales_file": (BytesIO(content.encode("utf-8")), filename)},
            content_type="multipart/form-data",
            follow_redirects=follow_redirects,
        )

    def test_index_and_upload_validation(self):
        response = self.client.get("/")
        self.assertEqual(response.status_code, 200)
        self.assertIn("Продажи становятся".encode(), response.data)
        response = self.upload(filename="sales.txt")
        self.assertEqual(response.status_code, 400)
        self.assertIn("Неподдерживаемый формат".encode(), response.data)

    def test_complete_upload_dashboard_filter_and_export(self):
        response = self.upload(follow_redirects=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("Панель продаж".encode(), response.data)
        self.assertIn("Прогноз спроса".encode(), response.data)

        empty = self.client.get("/dashboard?product=Несуществующий")
        self.assertEqual(empty.status_code, 200)
        self.assertIn("Нет данных для отображения".encode(), empty.data)

        export = self.client.get("/export?date_from=2025-02-01")
        self.assertEqual(export.status_code, 200)
        self.assertEqual(
            export.mimetype,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(export.data), read_only=True)
        self.assertEqual(workbook.sheetnames, ["Показатели", "Продажи", "По товарам", "По месяцам", "Прогноз"])
        self.assertEqual(workbook["Продажи"].max_row, 3)

    def test_invalid_file_does_not_remain_in_uploads(self):
        response = self.upload("Дата;Товар\n01.01.2025;Молоко\n")
        self.assertEqual(response.status_code, 400)
        self.assertEqual(list(self.uploads.iterdir()), [])


if __name__ == "__main__":
    unittest.main()
