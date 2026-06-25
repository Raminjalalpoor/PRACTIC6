from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="10233F")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ACCENT_FILL = PatternFill("solid", fgColor="E6F7F1")
CURRENCY_FORMAT = '#,##0.00 [$₽-ru-RU]'
NUMBER_FORMAT = '#,##0.00'


def _safe_text(value):
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _safe_frame(frame: pd.DataFrame) -> pd.DataFrame:
    result = frame.copy()
    for column in result.select_dtypes(include="object").columns:
        result[column] = result[column].map(_safe_text)
    return result


def _style_sheet(sheet, currency_columns=(), number_columns=()):
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 24

    headers = {cell.value: cell.column for cell in sheet[1]}
    for name in currency_columns:
        if name in headers:
            for cell in sheet[get_column_letter(headers[name])][1:]:
                cell.number_format = CURRENCY_FORMAT
    for name in number_columns:
        if name in headers:
            for cell in sheet[get_column_letter(headers[name])][1:]:
                cell.number_format = NUMBER_FORMAT

    for column_cells in sheet.columns:
        values = [len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells]
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(values, default=10) + 3, 45)


def build_excel_report(frame: pd.DataFrame, analysis, filters: dict) -> BytesIO:
    metrics = pd.DataFrame(
        [
            ("Общая выручка", analysis.metrics["revenue"]),
            ("Продано единиц", analysis.metrics["units"]),
            ("Количество записей", analysis.metrics["records"]),
            ("Уникальных товаров", analysis.metrics["unique_products"]),
            ("Средняя сумма операции", analysis.metrics["average_operation"]),
            ("Популярный товар", _safe_text(analysis.metrics["popular_product"])),
            ("Лидер по выручке", _safe_text(analysis.metrics["revenue_leader"])),
            ("Начало периода", filters.get("date_from") or "Все даты"),
            ("Конец периода", filters.get("date_to") or "Все даты"),
            ("Фильтр по товару", _safe_text(filters.get("product") or "Все товары")),
        ],
        columns=["Показатель", "Значение"],
    )

    sales = _safe_frame(frame[["Дата", "Товар", "Категория", "Количество", "Цена", "Выручка"]])
    products = _safe_frame(analysis.products)
    monthly = analysis.monthly[["Месяц", "Количество", "Выручка"]].copy()
    forecast = analysis.forecast["products"].copy()
    if analysis.forecast["available"]:
        forecast = forecast.dropna(axis=1, how="all")
        total_values = {column: None for column in forecast.columns}
        total_values.update({"Товар": "ВСЕГО", "Прогноз": analysis.forecast["total"]})
        forecast.index = forecast.index + 1
        forecast.loc[0] = total_values
        forecast = forecast.sort_index().reset_index(drop=True)
    forecast = _safe_frame(forecast)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl", datetime_format="DD.MM.YYYY") as writer:
        metrics.to_excel(writer, sheet_name="Показатели", index=False)
        sales.to_excel(writer, sheet_name="Продажи", index=False)
        products.to_excel(writer, sheet_name="По товарам", index=False)
        monthly.to_excel(writer, sheet_name="По месяцам", index=False)
        forecast.to_excel(writer, sheet_name="Прогноз", index=False)

        workbook = writer.book
        _style_sheet(workbook["Показатели"])
        for row in range(2, 7):
            workbook["Показатели"].cell(row=row, column=1).fill = ACCENT_FILL
        workbook["Показатели"]["B2"].number_format = CURRENCY_FORMAT
        workbook["Показатели"]["B6"].number_format = CURRENCY_FORMAT

        _style_sheet(workbook["Продажи"], currency_columns=("Цена", "Выручка"), number_columns=("Количество",))
        for cell in workbook["Продажи"]["A"][1:]:
            cell.number_format = "DD.MM.YYYY"
        _style_sheet(workbook["По товарам"], currency_columns=("Выручка",), number_columns=("Количество",))
        _style_sheet(workbook["По месяцам"], currency_columns=("Выручка",), number_columns=("Количество",))
        _style_sheet(workbook["Прогноз"], number_columns=("Прогноз", "Последний месяц", "Изменение"))

    output.seek(0)
    return output
